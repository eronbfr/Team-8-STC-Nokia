"""
Merge spreadsheet data: combines the existing (old) xlsx from git history
with the newly uploaded xlsx so that no data is lost.

Logic:
- The OLD workbook (from git history) is used as the base, preserving the
  full structure (all members, all dates, formatting).
- For each member present in the new upload, only cells with a non-zero
  value overwrite the corresponding cell in the old workbook.
- Members and dates are matched by name/date value.

Run by the CI workflow before dashboard_steps.py.
"""

import io
import os
import subprocess
import sys
from datetime import datetime

import openpyxl


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_NAME = 'step-tracking_Team8.xlsx'
XLSX_PATH = os.path.join(BASE_DIR, XLSX_NAME)


def _get_old_xlsx_bytes():
    """Return the previous version of the xlsx from git history, or None."""
    try:
        result = subprocess.run(
            ['git', 'show', 'HEAD~1:' + XLSX_NAME],
            capture_output=True,
            cwd=BASE_DIR,
        )
        if result.returncode == 0 and result.stdout:
            return result.stdout
    except Exception:
        pass
    return None


def _read_data(wb):
    """Read member step data from a workbook.

    Returns:
        dates: list of date values (datetime or str) from row 2
        members: dict of {name: {col_index: steps_value}}
    """
    # Try known sheet names, fall back to first sheet
    for name in ('Team 8', 'Team X'):
        if name in wb.sheetnames:
            ws = wb[name]
            break
    else:
        ws = wb[wb.sheetnames[0]]

    # Read dates from row 2, columns E (5) onwards
    dates = {}
    for col in range(5, ws.max_column + 1):
        val = ws.cell(row=2, column=col).value
        if val is not None:
            # Normalize to string for matching
            if isinstance(val, datetime):
                key = val.strftime('%Y-%m-%d')
            else:
                key = str(val)
            dates[col] = (key, val)

    # Read member rows (3–12)
    members = {}
    for row in range(3, min(13, ws.max_row + 1)):
        name = ws.cell(row=row, column=2).value
        if not name:
            continue
        name = str(name).strip()
        steps = {}
        for col, (date_key, _) in dates.items():
            val = ws.cell(row=row, column=col).value
            if val and isinstance(val, (int, float)):
                steps[date_key] = val
        members[name] = steps

    return dates, members


def merge():
    """Merge new upload into the existing spreadsheet, preserving its structure.

    The OLD workbook (from git history) is used as the base so that the full
    structure (all members, dates, formatting) is always preserved.  Only cells
    where the newly-uploaded file contains a non-zero value are overwritten.

    Example: the spreadsheet has 10 members.  Humberto uploads a file that
    only contains his own row → only Humberto's cells are updated; every
    other member's data remains untouched.
    """
    if not os.path.exists(XLSX_PATH):
        print("merge_xlsx: No xlsx file found, nothing to merge.")
        return

    old_bytes = _get_old_xlsx_bytes()
    if old_bytes is None:
        print("merge_xlsx: No previous version in git history, skipping merge.")
        return

    # Load OLD workbook as the BASE (preserves structure, all members, formatting)
    try:
        old_wb = openpyxl.load_workbook(io.BytesIO(old_bytes))
    except Exception as exc:
        print(f"merge_xlsx: Could not read old xlsx: {exc}")
        return

    # Load NEW (uploaded) workbook just to read data values
    try:
        new_wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    except Exception as exc:
        print(f"merge_xlsx: Could not read new xlsx: {exc}")
        old_wb.close()
        return

    _, new_members = _read_data(new_wb)

    # Get the worksheet from the OLD workbook (the base)
    for name in ('Team 8', 'Team X'):
        if name in old_wb.sheetnames:
            old_ws = old_wb[name]
            break
    else:
        old_ws = old_wb[old_wb.sheetnames[0]]

    # Build date_key → column index map for the OLD workbook
    old_date_to_col = {}
    for col in range(5, old_ws.max_column + 1):
        val = old_ws.cell(row=2, column=col).value
        if val is not None:
            key = val.strftime('%Y-%m-%d') if isinstance(val, datetime) else str(val)
            old_date_to_col[key] = col

    # Build member name → row index map for the OLD workbook
    old_name_to_row = {}
    for row in range(3, min(13, old_ws.max_row + 1)):
        name = old_ws.cell(row=row, column=2).value
        if name:
            old_name_to_row[str(name).strip()] = row

    updated_count = 0

    # For each member present in the NEW upload that also exists in the old
    # spreadsheet, overwrite only non-zero cells.
    for member_name, new_steps in new_members.items():
        if member_name not in old_name_to_row:
            continue  # member not in old sheet, skip
        row = old_name_to_row[member_name]

        for date_key, new_val in new_steps.items():
            if date_key not in old_date_to_col:
                continue  # date column not in old sheet
            col = old_date_to_col[date_key]

            # Only overwrite when the new upload provides a real value
            if new_val and new_val != 0:
                old_ws.cell(row=row, column=col).value = new_val
                updated_count += 1

    # Always save old workbook as result to guarantee structure is preserved
    old_wb.save(XLSX_PATH)

    if updated_count > 0:
        print(f"merge_xlsx: Updated {updated_count} cell(s) from new upload into existing spreadsheet.")
    else:
        print("merge_xlsx: No new data to merge; existing spreadsheet structure preserved.")

    old_wb.close()
    new_wb.close()


if __name__ == '__main__':
    merge()
