"""
Step-Tracking Dashboard Generator - Team 8
Reads the Excel file and generates a professional HTML dashboard.
Run: python dashboard_steps.py
Opens automatically in browser.
"""

import openpyxl
import json
import os
import webbrowser
import random
from datetime import datetime, timedelta

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
# Try multiple possible file names
_EXCEL_CANDIDATES = [
    'step-tracking_Team8.xlsx',
    'step-tracking 2026 - Team 8.xlsx',
    'step-tracking_Team 8.xlsx',
    'step-tracking_Team 8.csv',
]
EXCEL_PATH = None
for _name in _EXCEL_CANDIDATES:
    _path = os.path.join(BASE_DIR, _name)
    if os.path.exists(_path):
        EXCEL_PATH = _path
        break
OUTPUT_PATH = os.path.join(BASE_DIR, 'index.html')

# Daily goal per person
DAILY_GOAL = 10000
CHALLENGE_GOAL_TOTAL = DAILY_GOAL * 39  # 39 days

def read_excel_data():
    if not EXCEL_PATH:
        print("⚠️  No Excel file found. Using demo data.")
        # Return default structure for demo
        from datetime import timedelta
        dates = [(datetime(2026, 4, 7) + timedelta(days=i)).strftime('%Y-%m-%d') for i in range(39)]
        names = ['Juan Carlos Moran', 'Humberto Silva', 'Rocío Pérez', 'Ruben Salomoni',
                 'Eron Netto', 'Lizardo Ortiz Corzo', 'Alexis Rodriguez',
                 'Cynthia Villena', 'Ramiro Campos', 'Jessé Pereira Galvez']
        members = [{'name': n, 'daily_steps': [0]*39, 'total': 0} for n in names]
        return dates, members, True

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    # Try known sheet names, fall back to first sheet
    for _sheet_name in ('Team 8', 'Team X'):
        if _sheet_name in wb.sheetnames:
            ws = wb[_sheet_name]
            break
    else:
        ws = wb[wb.sheetnames[0]]

    # Read dates from row 2, starting at column E (5).
    # Walk the entire row instead of a hard-coded range so the dashboard
    # adapts gracefully if the spreadsheet ever grows beyond the original
    # 39-day window.  Track each (column, date) pair so member rows are read
    # from the *exact* columns that contained a date — preventing
    # mis-alignment when a date cell is blank.
    dates = []
    date_columns = []
    last_col = ws.max_column or 5
    for col in range(5, last_col + 1):
        val = ws.cell(row=2, column=col).value
        if val is None:
            continue
        if isinstance(val, datetime):
            dates.append(val.strftime('%Y-%m-%d'))
            date_columns.append(col)
        else:
            text = str(val).strip()
            if text:
                dates.append(text)
                date_columns.append(col)

    # Read members and their daily steps (rows 3–12 contain the team).
    members = []
    for row in range(3, 13):
        name = ws.cell(row=row, column=2).value
        if not name:
            continue
        daily_steps = []
        for col in date_columns:
            val = ws.cell(row=row, column=col).value
            # Only keep numeric values; treat blanks/strings/formulas as 0.
            if isinstance(val, bool) or not isinstance(val, (int, float)):
                daily_steps.append(0)
            else:
                daily_steps.append(val)
        members.append({
            'name': str(name).strip(),
            'daily_steps': daily_steps,
            'total': sum(daily_steps)
        })

    # Check if all data is zero (demo mode)
    all_zero = all(m['total'] == 0 for m in members)

    return dates, members, all_zero


def generate_demo_data(dates, members):
    """Generate realistic demo data when the spreadsheet is empty."""
    random.seed(42)
    base_ranges = [
        (7000, 14000), (8000, 15000), (6000, 12000), (9000, 16000), (7500, 13500),
        (6500, 11500), (8500, 14500), (7000, 13000), (9500, 15500), (8000, 14000)
    ]
    for i, member in enumerate(members):
        low, high = base_ranges[i % len(base_ranges)]
        daily = []
        for j, d in enumerate(dates):
            dt = datetime.strptime(d, '%Y-%m-%d')
            # Weekends slightly lower
            if dt.weekday() >= 5:
                steps = random.randint(int(low * 0.6), int(high * 0.8))
            else:
                steps = random.randint(low, high)
            # Future dates get 0
            if dt.date() > datetime.now().date():
                steps = 0
            daily.append(steps)
        member['daily_steps'] = daily
        member['total'] = sum(daily)
    return members


def generate_html(dates, members, is_demo):
    date_labels = []
    for d in dates:
        dt = datetime.strptime(d, '%Y-%m-%d')
        date_labels.append(dt.strftime('%b %d'))

    short_labels = []
    for d in dates:
        dt = datetime.strptime(d, '%Y-%m-%d')
        short_labels.append(dt.strftime('%d/%m'))

    # Compute week numbers
    week_data = {}
    for i, d in enumerate(dates):
        dt = datetime.strptime(d, '%Y-%m-%d')
        week_key = f"Week {dt.isocalendar()[1]}"
        if week_key not in week_data:
            week_data[week_key] = {}
        for m in members:
            if m['name'] not in week_data[week_key]:
                week_data[week_key][m['name']] = 0
            week_data[week_key][m['name']] += m['daily_steps'][i]

    # Compute stats
    total_team = sum(m['total'] for m in members)
    days_elapsed = 0
    for d in dates:
        dt = datetime.strptime(d, '%Y-%m-%d')
        if dt.date() <= datetime.now().date():
            days_elapsed += 1
    days_total = len(dates)

    avg_per_day_team = total_team / max(days_elapsed, 1)
    sorted_members = sorted(members, key=lambda x: x['total'], reverse=True)
    top_performer = sorted_members[0] if sorted_members else None

    # Best single day (only count numeric values & elapsed days)
    best_day_val = 0
    best_day_name = ""
    best_day_date = ""
    for m in members:
        for i, s in enumerate(m['daily_steps']):
            if not isinstance(s, (int, float)) or isinstance(s, bool):
                continue
            if s > best_day_val:
                best_day_val = s
                best_day_name = m['name'].split()[0]
                best_day_date = date_labels[i] if i < len(date_labels) else ""

    # ----- Health metrics ------------------------------------------------
    # Conversion factors based on widely-cited public-health references
    # (e.g. CDC / NIH / WHO step-tracking guidance):
    #   * 1 step ≈ 0.762 m (average adult stride)
    #   * 1 step ≈ 0.04 kcal burned (moderate-pace walking)
    #   * 1 active minute ≈ 100 brisk steps
    #   * 1 floor climbed ≈ 2,000 steps equivalent activity
    METERS_PER_STEP = 0.762
    KCAL_PER_STEP = 0.04
    STEPS_PER_ACTIVE_MIN = 100
    STEPS_PER_FLOOR = 2000

    total_distance_km = total_team * METERS_PER_STEP / 1000.0
    total_calories = total_team * KCAL_PER_STEP
    total_active_minutes = total_team / STEPS_PER_ACTIVE_MIN
    total_floors = total_team / STEPS_PER_FLOOR

    # Fun real-world equivalences
    marathons = total_distance_km / 42.195
    earth_pct = (total_distance_km / 40075.0) * 100  # Earth's circumference
    eiffel_climbs = total_floors / 1665 * 100  # Eiffel Tower has ~1665 steps -> floors approx

    # ----- Goal achievement & projection metrics -------------------------
    person_days_evaluated = 0
    person_days_hit_goal = 0
    for m in members:
        for i in range(min(days_elapsed, len(m['daily_steps']))):
            person_days_evaluated += 1
            if m['daily_steps'][i] >= DAILY_GOAL:
                person_days_hit_goal += 1
    goal_hit_rate = (person_days_hit_goal / person_days_evaluated * 100) if person_days_evaluated else 0

    # Project total team steps for the full challenge at the current pace
    if days_elapsed > 0:
        projected_total = total_team / days_elapsed * days_total
    else:
        projected_total = 0
    challenge_goal_team = DAILY_GOAL * len(members) * days_total

    # Most consistent member (lowest coefficient of variation across elapsed days)
    most_consistent_name = "-"
    most_consistent_avg = 0
    if days_elapsed >= 3:
        best_cv = None
        for m in members:
            elapsed_steps = m['daily_steps'][:days_elapsed]
            non_zero = [s for s in elapsed_steps if s > 0]
            if len(non_zero) < 3:
                continue
            mean_v = sum(non_zero) / len(non_zero)
            if mean_v <= 0:
                continue
            variance = sum((s - mean_v) ** 2 for s in non_zero) / len(non_zero)
            cv = (variance ** 0.5) / mean_v
            if best_cv is None or cv < best_cv:
                best_cv = cv
                most_consistent_name = m['name'].split()[0]
                most_consistent_avg = mean_v

    # Per-member goal-hit days & longest streak (used in leaderboard)
    for m in members:
        goal_days = 0
        current_streak = 0
        longest_streak = 0
        for i in range(min(days_elapsed, len(m['daily_steps']))):
            if m['daily_steps'][i] >= DAILY_GOAL:
                goal_days += 1
                current_streak += 1
                if current_streak > longest_streak:
                    longest_streak = current_streak
            else:
                current_streak = 0
        m['goal_days'] = goal_days
        m['longest_streak'] = longest_streak

    # Average steps grouped by day of the week (for weekday-pattern chart)
    weekday_totals = [0] * 7
    weekday_counts = [0] * 7
    for i, d in enumerate(dates):
        if i >= days_elapsed:
            break
        wd = datetime.strptime(d, '%Y-%m-%d').weekday()  # 0=Mon
        for m in members:
            if i < len(m['daily_steps']):
                weekday_totals[wd] += m['daily_steps'][i]
                weekday_counts[wd] += 1
    weekday_avg = [
        (weekday_totals[i] / weekday_counts[i]) if weekday_counts[i] else 0
        for i in range(7)
    ]
    weekday_labels = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']

    # Member colors (Nokia vibrant palette)
    colors = [
        '#124191', '#009DE0', '#FF3D71', '#00C48C', '#FFA940',
        '#7B61FF', '#FF6F91', '#0080C0', '#F7B731', '#26D07C'
    ]

    member_data_json = json.dumps([{
        'name': m['name'],
        'total': m['total'],
        'daily': m['daily_steps'],
        'avg': round(m['total'] / max(days_elapsed, 1)),
        'goalDays': m.get('goal_days', 0),
        'streak': m.get('longest_streak', 0),
        'color': colors[i % len(colors)]
    } for i, m in enumerate(members)])

    date_labels_json = json.dumps(date_labels)
    short_labels_json = json.dumps(short_labels)
    week_labels = list(week_data.keys())
    week_labels_json = json.dumps(week_labels)

    # Weekly totals per member
    weekly_member_data = {}
    for wk in week_labels:
        for m in members:
            if m['name'] not in weekly_member_data:
                weekly_member_data[m['name']] = []
            weekly_member_data[m['name']].append(week_data[wk].get(m['name'], 0))

    weekly_member_json = json.dumps(weekly_member_data)
    weekday_avg_json = json.dumps(weekday_avg)
    weekday_labels_json = json.dumps(weekday_labels)

    demo_banner = ""
    if is_demo:
        demo_banner = """
        <div class="demo-banner">
            <span class="demo-icon">🎯</span>
            <span>DEMO MODE — The data below is simulated. Upload your Excel spreadsheet to see real data.</span>
        </div>"""

    html = f"""<!DOCTYPE html>
<html lang="en-US">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>🏃 Step Tracking Dashboard — Team 8 | 2026</title>
    <script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.7/dist/chart.umd.min.js"></script>
    <script src="https://cdn.jsdelivr.net/npm/chartjs-plugin-datalabels@2.2.0/dist/chartjs-plugin-datalabels.min.js"></script>
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800;900&family=JetBrains+Mono:wght@400;500;600&display=swap" rel="stylesheet">
    <style>
        :root {{
            --bg-primary: #F0F4F8;
            --bg-secondary: #E8EDF3;
            --bg-card: #FFFFFF;
            --bg-card-hover: #F7F9FC;
            --accent-1: #124191;
            --accent-2: #009DE0;
            --accent-3: #FF3D71;
            --accent-4: #FFA940;
            --accent-5: #00C48C;
            --text-primary: #1A2138;
            --text-secondary: #4A5568;
            --text-muted: #8892A4;
            --border: #E2E8F0;
            --gradient-1: linear-gradient(135deg, #124191 0%, #1B6AC9 100%);
            --gradient-2: linear-gradient(135deg, #009DE0 0%, #0070C0 100%);
            --gradient-3: linear-gradient(135deg, #FF3D71 0%, #E0245E 100%);
            --gradient-4: linear-gradient(135deg, #FFA940 0%, #FF7A00 100%);
            --gradient-5: linear-gradient(135deg, #00C48C 0%, #009E73 100%);
            --shadow: 0 4px 24px rgba(18,65,145,0.08);
            --shadow-glow: 0 0 40px rgba(18,65,145,0.06);
        }}

        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}

        body {{
            font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif;
            background: var(--bg-primary);
            color: var(--text-primary);
            min-height: 100vh;
            overflow-x: hidden;
        }}

        /* Animated background */
        body::before {{
            content: '';
            position: fixed;
            top: 0;
            left: 0;
            width: 100%;
            height: 100%;
            background: 
                radial-gradient(ellipse at 20% 50%, rgba(18,65,145,0.04) 0%, transparent 50%),
                radial-gradient(ellipse at 80% 20%, rgba(0,157,224,0.04) 0%, transparent 50%),
                radial-gradient(ellipse at 50% 80%, rgba(0,196,140,0.03) 0%, transparent 50%);
            pointer-events: none;
            z-index: 0;
        }}

        .dashboard {{
            position: relative;
            z-index: 1;
            max-width: 1600px;
            margin: 0 auto;
            padding: 24px 32px 48px;
        }}

        /* Header */
        .header {{
            display: flex;
            align-items: center;
            justify-content: space-between;
            padding: 28px 36px;
            background: var(--bg-card);
            border-radius: 0 0 20px 20px;
            border: 1px solid var(--border);
            border-top: none;
            margin-bottom: 28px;
            box-shadow: var(--shadow);
            position: sticky;
            top: 0;
            z-index: 100;
            overflow: hidden;
        }}

        .header::before {{
            content: '';
            position: absolute;
            top: 0;
            left: 0;
            right: 0;
            height: 3px;
            background: linear-gradient(90deg, #124191, #009DE0, #FF3D71, #FFA940, #00C48C);
        }}

        .header-left h1 {{
            font-size: 28px;
            font-weight: 800;
            letter-spacing: -0.5px;
            background: linear-gradient(135deg, #124191 0%, #1B6AC9 100%);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
        }}

        .header-left p {{
            color: var(--text-secondary);
            font-size: 14px;
            margin-top: 4px;
            font-weight: 400;
        }}

        .header-right {{
            display: flex;
            align-items: center;
            gap: 16px;
        }}

        .header-badge {{
            display: flex;
            align-items: center;
            gap: 8px;
            padding: 8px 16px;
            background: rgba(18,65,145,0.08);
            border: 1px solid rgba(18,65,145,0.2);
            border-radius: 12px;
            font-size: 13px;
            font-weight: 600;
            color: #124191;
        }}

        .pulse-dot {{
            width: 8px;
            height: 8px;
            background: #00C48C;
            border-radius: 50%;
            animation: pulse 2s infinite;
        }}

        @keyframes pulse {{
            0%, 100% {{ opacity: 1; transform: scale(1); }}
            50% {{ opacity: 0.5; transform: scale(1.3); }}
        }}

        /* Demo Banner */
        .demo-banner {{
            display: flex;
            align-items: center;
            gap: 12px;
            padding: 14px 24px;
            background: linear-gradient(135deg, rgba(255,169,64,0.1) 0%, rgba(255,122,0,0.05) 100%);
            border: 1px solid rgba(255,169,64,0.35);
            border-radius: 14px;
            margin-bottom: 24px;
            font-size: 13px;
            color: #B87300;
            font-weight: 500;
        }}

        .demo-icon {{
            font-size: 20px;
        }}

        /* KPI Cards */
        .kpi-grid {{
            display: grid;
            grid-template-columns: repeat(5, 1fr);
            gap: 20px;
            margin-bottom: 28px;
        }}

        .kpi-card {{
            background: var(--bg-card);
            border-radius: 18px;
            padding: 24px;
            border: 1px solid var(--border);
            box-shadow: var(--shadow);
            position: relative;
            overflow: hidden;
            transition: transform 0.3s ease, box-shadow 0.3s ease;
        }}

        .kpi-card:hover {{
            transform: translateY(-4px);
            box-shadow: 0 12px 40px rgba(18,65,145,0.12);
        }}

        .kpi-card::before {{
            content: '';
            position: absolute;
            top: 0;
            left: 0;
            right: 0;
            height: 3px;
        }}

        .kpi-card:nth-child(1)::before {{ background: var(--gradient-1); }}
        .kpi-card:nth-child(2)::before {{ background: var(--gradient-2); }}
        .kpi-card:nth-child(3)::before {{ background: var(--gradient-3); }}
        .kpi-card:nth-child(4)::before {{ background: var(--gradient-4); }}
        .kpi-card:nth-child(5)::before {{ background: var(--gradient-5); }}

        .kpi-icon {{
            width: 44px;
            height: 44px;
            border-radius: 12px;
            display: flex;
            align-items: center;
            justify-content: center;
            font-size: 22px;
            margin-bottom: 16px;
        }}

        .kpi-card:nth-child(1) .kpi-icon {{ background: rgba(18,65,145,0.1); }}
        .kpi-card:nth-child(2) .kpi-icon {{ background: rgba(0,157,224,0.1); }}
        .kpi-card:nth-child(3) .kpi-icon {{ background: rgba(255,61,113,0.1); }}
        .kpi-card:nth-child(4) .kpi-icon {{ background: rgba(255,169,64,0.12); }}
        .kpi-card:nth-child(5) .kpi-icon {{ background: rgba(0,196,140,0.1); }}

        .kpi-label {{
            font-size: 12px;
            font-weight: 600;
            text-transform: uppercase;
            letter-spacing: 1px;
            color: var(--text-muted);
            margin-bottom: 8px;
        }}

        .kpi-value {{
            font-size: 30px;
            font-weight: 800;
            font-family: 'JetBrains Mono', monospace;
            letter-spacing: -1px;
        }}

        .kpi-card:nth-child(1) .kpi-value {{ color: #124191; }}
        .kpi-card:nth-child(2) .kpi-value {{ color: #0080C0; }}
        .kpi-card:nth-child(3) .kpi-value {{ color: #E0245E; }}
        .kpi-card:nth-child(4) .kpi-value {{ color: #D48400; }}
        .kpi-card:nth-child(5) .kpi-value {{ color: #009E73; }}

        .kpi-sub {{
            font-size: 12px;
            color: var(--text-muted);
            margin-top: 4px;
        }}

        /* Health Impact band ------------------------------------------- */
        .health-band {{
            background: var(--bg-card);
            border-radius: 18px;
            padding: 22px 26px;
            border: 1px solid var(--border);
            box-shadow: var(--shadow);
            margin-bottom: 28px;
            position: relative;
            overflow: hidden;
        }}
        .health-band::before {{
            content: '';
            position: absolute;
            top: 0; left: 0; right: 0;
            height: 3px;
            background: linear-gradient(90deg, #00C48C, #009DE0, #124191);
        }}
        .health-band-title {{
            font-size: 13px;
            font-weight: 700;
            text-transform: uppercase;
            letter-spacing: 1px;
            color: var(--text-muted);
            margin-bottom: 14px;
            display: flex;
            align-items: center;
            gap: 10px;
        }}
        .health-band-title .health-tag {{
            background: linear-gradient(135deg, #00C48C 0%, #009E73 100%);
            color: #fff;
            padding: 3px 10px;
            border-radius: 8px;
            font-size: 10px;
            letter-spacing: 0.5px;
        }}
        .health-grid {{
            display: grid;
            grid-template-columns: repeat(4, 1fr);
            gap: 18px;
        }}
        .health-item {{
            display: flex;
            align-items: center;
            gap: 14px;
            padding: 14px 16px;
            background: var(--bg-secondary);
            border-radius: 14px;
            border: 1px solid transparent;
            transition: all 0.3s ease;
        }}
        .health-item:hover {{
            transform: translateY(-2px);
            border-color: var(--border);
            background: var(--bg-card-hover);
        }}
        .health-item-icon {{
            width: 42px;
            height: 42px;
            border-radius: 12px;
            display: flex;
            align-items: center;
            justify-content: center;
            font-size: 20px;
            flex-shrink: 0;
        }}
        .health-item:nth-child(1) .health-item-icon {{ background: rgba(0,196,140,0.15); }}
        .health-item:nth-child(2) .health-item-icon {{ background: rgba(255,61,113,0.12); }}
        .health-item:nth-child(3) .health-item-icon {{ background: rgba(0,157,224,0.12); }}
        .health-item:nth-child(4) .health-item-icon {{ background: rgba(255,169,64,0.14); }}
        .health-item-body {{
            min-width: 0;
        }}
        .health-item-value {{
            font-family: 'JetBrains Mono', monospace;
            font-size: 20px;
            font-weight: 800;
            color: var(--text-primary);
            letter-spacing: -0.5px;
            line-height: 1.1;
        }}
        .health-item-label {{
            font-size: 11px;
            font-weight: 600;
            text-transform: uppercase;
            letter-spacing: 0.6px;
            color: var(--text-muted);
            margin-top: 4px;
        }}
        .health-item-sub {{
            font-size: 11px;
            color: var(--text-secondary);
            margin-top: 2px;
        }}

        /* Insight cards (consistency / projection) --------------------- */
        .insight-grid {{
            display: grid;
            grid-template-columns: repeat(3, 1fr);
            gap: 20px;
            margin-bottom: 28px;
        }}
        .insight-card {{
            background: var(--bg-card);
            border-radius: 16px;
            padding: 20px 22px;
            border: 1px solid var(--border);
            box-shadow: var(--shadow);
            display: flex;
            align-items: center;
            gap: 16px;
        }}
        .insight-icon {{
            width: 48px;
            height: 48px;
            border-radius: 14px;
            display: flex;
            align-items: center;
            justify-content: center;
            font-size: 22px;
            flex-shrink: 0;
        }}
        .insight-card:nth-child(1) .insight-icon {{ background: rgba(18,65,145,0.10); }}
        .insight-card:nth-child(2) .insight-icon {{ background: rgba(0,157,224,0.10); }}
        .insight-card:nth-child(3) .insight-icon {{ background: rgba(0,196,140,0.10); }}
        .insight-body {{ flex: 1; min-width: 0; }}
        .insight-label {{
            font-size: 11px;
            font-weight: 700;
            text-transform: uppercase;
            letter-spacing: 0.8px;
            color: var(--text-muted);
        }}
        .insight-value {{
            font-family: 'JetBrains Mono', monospace;
            font-size: 22px;
            font-weight: 800;
            color: var(--text-primary);
            margin-top: 4px;
        }}
        .insight-sub {{
            font-size: 11px;
            color: var(--text-secondary);
            margin-top: 4px;
        }}

        /* Leaderboard chips */
        .leader-chips {{
            display: flex;
            gap: 6px;
            margin-top: 4px;
            flex-wrap: wrap;
        }}
        .leader-chip {{
            font-size: 10px;
            font-weight: 600;
            padding: 2px 7px;
            border-radius: 999px;
            background: rgba(0,196,140,0.12);
            color: #009E73;
            white-space: nowrap;
        }}
        .leader-chip.streak {{
            background: rgba(255,169,64,0.14);
            color: #B87300;
        }}

        /* Charts Grid */
        .charts-grid {{
            display: grid;
            grid-template-columns: 2fr 1fr;
            gap: 24px;
            margin-bottom: 28px;
        }}

        .chart-card {{
            background: var(--bg-card);
            border-radius: 18px;
            padding: 28px;
            border: 1px solid var(--border);
            box-shadow: var(--shadow);
        }}

        .chart-title {{
            font-size: 16px;
            font-weight: 700;
            margin-bottom: 20px;
            display: flex;
            align-items: center;
            gap: 10px;
        }}

        .chart-title-icon {{
            width: 32px;
            height: 32px;
            border-radius: 8px;
            display: flex;
            align-items: center;
            justify-content: center;
            font-size: 16px;
        }}

        .chart-container {{
            position: relative;
            width: 100%;
        }}

        /* Ranking / Leaderboard */
        .leaderboard {{
            display: flex;
            flex-direction: column;
            gap: 10px;
        }}

        .leader-row {{
            display: flex;
            align-items: center;
            gap: 14px;
            padding: 12px 16px;
            border-radius: 14px;
            background: var(--bg-secondary);
            border: 1px solid transparent;
            transition: all 0.3s ease;
        }}

        .leader-row:hover {{
            background: var(--bg-card-hover);
            border-color: var(--border);
        }}

        .leader-row.gold {{
            background: linear-gradient(135deg, rgba(255,215,0,0.08) 0%, rgba(255,215,0,0.02) 100%);
            border: 1px solid rgba(255,215,0,0.2);
        }}

        .leader-row.silver {{
            background: linear-gradient(135deg, rgba(192,192,192,0.06) 0%, rgba(192,192,192,0.02) 100%);
            border: 1px solid rgba(192,192,192,0.15);
        }}

        .leader-row.bronze {{
            background: linear-gradient(135deg, rgba(205,127,50,0.06) 0%, rgba(205,127,50,0.02) 100%);
            border: 1px solid rgba(205,127,50,0.15);
        }}

        .leader-rank {{
            width: 32px;
            height: 32px;
            border-radius: 10px;
            display: flex;
            align-items: center;
            justify-content: center;
            font-size: 14px;
            font-weight: 800;
            flex-shrink: 0;
        }}

        .leader-row.gold .leader-rank {{
            background: linear-gradient(135deg, #FFD700, #FFA500);
            color: #000;
        }}

        .leader-row.silver .leader-rank {{
            background: linear-gradient(135deg, #C0C0C0, #A0A0A0);
            color: #000;
        }}

        .leader-row.bronze .leader-rank {{
            background: linear-gradient(135deg, #CD7F32, #A0522D);
            color: #fff;
        }}

        .leader-rank.other {{
            background: var(--bg-card);
            color: var(--text-muted);
            font-size: 13px;
        }}

        .leader-info {{
            flex: 1;
            min-width: 0;
        }}

        .leader-name {{
            font-size: 13px;
            font-weight: 600;
            white-space: nowrap;
            overflow: hidden;
            text-overflow: ellipsis;
        }}

        .leader-avg {{
            font-size: 11px;
            color: var(--text-muted);
            margin-top: 2px;
        }}

        .leader-steps {{
            font-family: 'JetBrains Mono', monospace;
            font-size: 14px;
            font-weight: 700;
            color: var(--accent-2);
            flex-shrink: 0;
        }}

        .leader-bar-wrapper {{
            width: 80px;
            height: 6px;
            background: rgba(18,65,145,0.06);
            border-radius: 3px;
            overflow: hidden;
            flex-shrink: 0;
        }}

        .leader-bar {{
            height: 100%;
            border-radius: 3px;
            transition: width 1s ease;
        }}

        /* Bottom Grid */
        .bottom-grid {{
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 24px;
            margin-bottom: 28px;
        }}

        /* Weekly Heatmap */
        .heatmap-grid {{
            display: grid;
            grid-template-columns: 120px repeat(7, 1fr);
            gap: 4px;
            font-size: 11px;
        }}

        .heatmap-header {{
            font-weight: 700;
            color: var(--text-muted);
            text-align: center;
            padding: 6px 2px;
            font-size: 10px;
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }}

        .heatmap-name {{
            font-size: 11px;
            font-weight: 600;
            color: var(--text-secondary);
            display: flex;
            align-items: center;
            padding-right: 8px;
            white-space: nowrap;
            overflow: hidden;
            text-overflow: ellipsis;
        }}

        .heatmap-cell {{
            border-radius: 6px;
            display: flex;
            align-items: center;
            justify-content: center;
            font-family: 'JetBrains Mono', monospace;
            font-size: 10px;
            font-weight: 600;
            padding: 8px 2px;
            transition: transform 0.2s ease;
        }}

        .heatmap-cell:hover {{
            transform: scale(1.15);
            z-index: 2;
        }}

        /* Progress Rings */
        .progress-container {{
            display: flex;
            align-items: center;
            justify-content: center;
            gap: 40px;
            padding: 20px 0;
        }}

        .ring-wrapper {{
            text-align: center;
        }}

        .ring-label {{
            font-size: 12px;
            font-weight: 600;
            color: var(--text-muted);
            margin-top: 12px;
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }}

        .ring-value {{
            font-size: 11px;
            color: var(--text-secondary);
            margin-top: 4px;
        }}

        /* Footer */
        .footer {{
            text-align: center;
            padding: 20px;
            color: var(--text-muted);
            font-size: 12px;
        }}

        .footer a {{
            color: var(--accent-1);
            text-decoration: none;
        }}

        /* Responsive */
        @media (max-width: 1200px) {{
            .kpi-grid {{ grid-template-columns: repeat(3, 1fr); }}
            .charts-grid {{ grid-template-columns: 1fr; }}
            .bottom-grid {{ grid-template-columns: 1fr; }}
            .health-grid {{ grid-template-columns: repeat(2, 1fr); }}
            .insight-grid {{ grid-template-columns: 1fr; }}
        }}

        @media (max-width: 768px) {{
            .kpi-grid {{ grid-template-columns: repeat(2, 1fr); }}
            .dashboard {{ padding: 12px; }}
            .header {{
                flex-direction: column;
                gap: 12px;
                text-align: center;
                padding: 16px 14px;
                border-radius: 0 0 16px 16px;
            }}
            .header-left h1 {{ font-size: 20px; }}
            .header-left p {{ font-size: 12px; }}
            .header-right {{
                flex-wrap: wrap;
                justify-content: center;
                gap: 8px;
            }}
            .header-badge {{ padding: 6px 10px; font-size: 11px; }}
            .upload-btn {{ padding: 6px 12px; font-size: 12px; }}
            .kpi-card {{ padding: 16px; border-radius: 14px; }}
            .kpi-value {{ font-size: 22px; }}
            .kpi-icon {{ width: 36px; height: 36px; font-size: 18px; margin-bottom: 10px; }}
            .chart-card {{ padding: 16px; border-radius: 14px; }}
            .chart-title {{ font-size: 14px; }}
            .tab-group {{ flex-wrap: wrap; }}
            .tab-btn {{ padding: 6px 12px; font-size: 11px; }}
            .leader-row {{ padding: 10px 12px; gap: 10px; }}
            .leader-name {{ font-size: 12px; }}
            .leader-steps {{ font-size: 12px; }}
            .leader-bar-wrapper {{ width: 50px; }}
            .heatmap-grid {{
                grid-template-columns: 80px repeat(7, 1fr);
                gap: 2px;
            }}
            .heatmap-name {{ font-size: 10px; }}
            .heatmap-cell {{ font-size: 9px; padding: 6px 1px; }}
            .demo-banner {{ font-size: 12px; padding: 10px 14px; }}
            .health-grid {{ grid-template-columns: repeat(2, 1fr); gap: 12px; }}
            .health-band {{ padding: 16px 14px; border-radius: 14px; }}
            .health-item {{ padding: 10px 12px; gap: 10px; }}
            .health-item-icon {{ width: 36px; height: 36px; font-size: 17px; border-radius: 10px; }}
            .health-item-value {{ font-size: 16px; }}
            .insight-card {{ padding: 14px 16px; border-radius: 14px; gap: 12px; }}
            .insight-icon {{ width: 40px; height: 40px; font-size: 18px; }}
            .insight-value {{ font-size: 18px; }}
        }}

        @media (max-width: 480px) {{
            .kpi-grid {{ grid-template-columns: 1fr; }}
            .dashboard {{ padding: 8px; }}
            .header {{
                padding: 14px 12px;
            }}
            .header-left h1 {{ font-size: 18px; }}
            .kpi-card {{ padding: 14px; }}
            .kpi-value {{ font-size: 20px; }}
            .kpi-label {{ font-size: 10px; }}
            .heatmap-grid {{
                grid-template-columns: 60px repeat(7, 1fr);
                gap: 1px;
            }}
            .heatmap-cell {{ font-size: 8px; padding: 4px 0; border-radius: 4px; }}
            .heatmap-name {{ font-size: 9px; }}
            .chart-container {{ min-height: 250px; }}
            .leader-bar-wrapper {{ display: none; }}
            .health-grid {{ grid-template-columns: 1fr; }}
        }}

        /* Animations */
        @keyframes fadeInUp {{
            from {{ opacity: 0; transform: translateY(20px); }}
            to {{ opacity: 1; transform: translateY(0); }}
        }}

        .animate {{
            animation: fadeInUp 0.6s ease forwards;
        }}

        .delay-1 {{ animation-delay: 0.1s; opacity: 0; }}
        .delay-2 {{ animation-delay: 0.2s; opacity: 0; }}
        .delay-3 {{ animation-delay: 0.3s; opacity: 0; }}
        .delay-4 {{ animation-delay: 0.4s; opacity: 0; }}
        .delay-5 {{ animation-delay: 0.5s; opacity: 0; }}

        /* Sparkline mini chart */
        .sparkline {{
            display: inline-block;
            vertical-align: middle;
        }}

        /* Scrollbar */
        ::-webkit-scrollbar {{
            width: 8px;
        }}
        ::-webkit-scrollbar-track {{
            background: var(--bg-primary);
        }}
        ::-webkit-scrollbar-thumb {{
            background: #C4CDD9;
            border-radius: 4px;
        }}
        ::-webkit-scrollbar-thumb:hover {{
            background: #9AA5B4;
        }}

        /* Tab buttons */
        .tab-group {{
            display: flex;
            gap: 8px;
            margin-bottom: 20px;
        }}

        .tab-btn {{
            padding: 8px 18px;
            border: 1px solid var(--border);
            background: transparent;
            color: var(--text-secondary);
            border-radius: 10px;
            cursor: pointer;
            font-size: 12px;
            font-weight: 600;
            font-family: 'Inter', sans-serif;
            transition: all 0.3s ease;
        }}

        .tab-btn.active {{
            background: var(--accent-1);
            border-color: var(--accent-1);
            color: #fff;
        }}

        .tab-btn:hover:not(.active) {{
            background: var(--bg-card-hover);
        }}

        /* Upload button */
        .upload-btn {{
            display: flex;
            align-items: center;
            gap: 6px;
            padding: 8px 16px;
            background: var(--gradient-1);
            border: none;
            border-radius: 12px;
            font-size: 13px;
            font-weight: 600;
            font-family: 'Inter', sans-serif;
            color: #fff;
            cursor: pointer;
            transition: all 0.3s ease;
            white-space: nowrap;
            text-decoration: none;
        }}

        .upload-btn:hover {{
            transform: translateY(-1px);
            box-shadow: 0 4px 16px rgba(18,65,145,0.25);
            color: #fff;
        }}

        .upload-btn:active {{
            transform: translateY(0);
        }}

        /* Upload modal overlay */
        .upload-overlay {{
            display: none;
            position: fixed;
            inset: 0;
            background: rgba(0,0,0,0.5);
            z-index: 9999;
            align-items: center;
            justify-content: center;
            backdrop-filter: blur(4px);
        }}
        .upload-overlay.active {{
            display: flex;
        }}
        .upload-modal {{
            background: #fff;
            border-radius: 16px;
            padding: 28px 32px;
            max-width: 420px;
            width: 90%;
            box-shadow: 0 20px 60px rgba(0,0,0,0.3);
            text-align: center;
            font-family: 'Inter', sans-serif;
        }}
        .upload-modal h3 {{
            margin: 0 0 8px;
            font-size: 18px;
            color: #1a202c;
        }}
        .upload-modal p {{
            margin: 0 0 18px;
            font-size: 13px;
            color: #64748b;
            line-height: 1.5;
        }}
        .upload-modal input[type="password"],
        .upload-modal input[type="text"] {{
            width: 100%;
            padding: 10px 14px;
            border: 1.5px solid #e2e8f0;
            border-radius: 10px;
            font-size: 13px;
            font-family: 'Inter', sans-serif;
            margin-bottom: 14px;
            box-sizing: border-box;
            outline: none;
            transition: border-color 0.2s;
        }}
        .upload-modal input:focus {{
            border-color: #124191;
        }}
        .upload-modal-btns {{
            display: flex;
            gap: 10px;
            justify-content: center;
        }}
        .upload-modal-btns button {{
            padding: 9px 20px;
            border-radius: 10px;
            border: none;
            font-size: 13px;
            font-weight: 600;
            font-family: 'Inter', sans-serif;
            cursor: pointer;
            transition: all 0.2s;
        }}
        .modal-btn-primary {{
            background: var(--gradient-1);
            color: #fff;
        }}
        .modal-btn-primary:hover {{
            box-shadow: 0 4px 12px rgba(18,65,145,0.3);
        }}
        .modal-btn-cancel {{
            background: #f1f5f9;
            color: #64748b;
        }}
        .modal-btn-cancel:hover {{
            background: #e2e8f0;
        }}
        /* Toast notification */
        .upload-toast {{
            position: fixed;
            top: 24px;
            right: 24px;
            z-index: 10000;
            padding: 14px 22px;
            border-radius: 12px;
            font-family: 'Inter', sans-serif;
            font-size: 13px;
            font-weight: 600;
            color: #fff;
            box-shadow: 0 8px 32px rgba(0,0,0,0.18);
            transform: translateX(120%);
            transition: transform 0.4s cubic-bezier(.4,0,.2,1);
            max-width: 360px;
        }}
        .upload-toast.show {{
            transform: translateX(0);
        }}
        .upload-toast.success {{
            background: linear-gradient(135deg, #10b981, #059669);
        }}
        .upload-toast.error {{
            background: linear-gradient(135deg, #ef4444, #dc2626);
        }}
        .upload-toast.info {{
            background: linear-gradient(135deg, #124191, #009DE0);
        }}
        /* Spinner inside button */
        .upload-spinner {{
            display: inline-block;
            width: 14px;
            height: 14px;
            border: 2px solid rgba(255,255,255,0.3);
            border-top-color: #fff;
            border-radius: 50%;
            animation: spin 0.6s linear infinite;
        }}
        @keyframes spin {{
            to {{ transform: rotate(360deg); }}
        }}

    </style>
</head>
<body>
    <div class="dashboard">
        <!-- Header -->
        <div class="header animate">
            <div class="header-left">
                <h1>🏃 Nokia SEAL TEAM 8</h1>
                <p>Step Tracking Challenge — Apr / May 2026 &nbsp;•&nbsp; Day {days_elapsed} of {days_total}</p>
            </div>
            <div class="header-right">
                <input type="file" id="xlsxFileInput" accept=".xlsx" style="display:none">
                <button class="upload-btn" id="uploadBtn" title="Enviar planilha ao repositório">
                    📤 Upload
                </button>
                
                <div class="header-badge">
                    <div class="pulse-dot"></div>
                    <span>{days_total - days_elapsed} days left</span>
                </div>
                <div class="header-badge" style="background: rgba(0,157,224,0.08); border-color: rgba(0,157,224,0.2);">
                    <span style="color: #009DE0;">👥 {len(members)} members</span>
                </div>
            </div>
        </div>

        {demo_banner}

        <!-- KPI Cards -->
        <div class="kpi-grid">
            <div class="kpi-card animate delay-1">
                <div class="kpi-icon">👣</div>
                <div class="kpi-label">Total Steps (Team)</div>
                <div class="kpi-value" id="kpi-total">{total_team:,.0f}</div>
                <div class="kpi-sub">Goal: {DAILY_GOAL * len(members) * days_total:,.0f}</div>
            </div>
            <div class="kpi-card animate delay-2">
                <div class="kpi-icon">📊</div>
                <div class="kpi-label">Daily Avg / Team</div>
                <div class="kpi-value" id="kpi-avg">{avg_per_day_team:,.0f}</div>
                <div class="kpi-sub">Daily goal: {DAILY_GOAL * len(members):,.0f}</div>
            </div>
            <div class="kpi-card animate delay-3">
                <div class="kpi-icon">🏆</div>
                <div class="kpi-label">Top Performer</div>
                <div class="kpi-value" style="font-size: 22px;">{top_performer['name'].split()[0] if top_performer else '-'}</div>
                <div class="kpi-sub">{top_performer['total']:,.0f} steps</div>
            </div>
            <div class="kpi-card animate delay-4">
                <div class="kpi-icon">🔥</div>
                <div class="kpi-label">Daily Record</div>
                <div class="kpi-value">{best_day_val:,.0f}</div>
                <div class="kpi-sub">{best_day_name} — {best_day_date}</div>
            </div>
            <div class="kpi-card animate delay-5">
                <div class="kpi-icon">🎯</div>
                <div class="kpi-label">Progress</div>
                <div class="kpi-value">{(total_team / max(DAILY_GOAL * len(members) * days_total, 1) * 100):.1f}%</div>
                <div class="kpi-sub">Day {days_elapsed} of {days_total}</div>
            </div>
        </div>

        <!-- Health Impact band -->
        <div class="health-band animate delay-2">
            <div class="health-band-title">
                <span class="health-tag">HEALTH</span>
                Real-world impact of every step taken by the team
            </div>
            <div class="health-grid">
                <div class="health-item">
                    <div class="health-item-icon">📏</div>
                    <div class="health-item-body">
                        <div class="health-item-value">{total_distance_km:,.1f} km</div>
                        <div class="health-item-label">Distance walked</div>
                        <div class="health-item-sub">≈ {marathons:.1f} marathons &nbsp;·&nbsp; {earth_pct:.2f}% around Earth</div>
                    </div>
                </div>
                <div class="health-item">
                    <div class="health-item-icon">🔥</div>
                    <div class="health-item-body">
                        <div class="health-item-value">{total_calories:,.0f} kcal</div>
                        <div class="health-item-label">Calories burned</div>
                        <div class="health-item-sub">≈ {(total_calories / 250):.0f} chocolate bars worth</div>
                    </div>
                </div>
                <div class="health-item">
                    <div class="health-item-icon">⏱️</div>
                    <div class="health-item-body">
                        <div class="health-item-value">{total_active_minutes:,.0f} min</div>
                        <div class="health-item-label">Active minutes</div>
                        <div class="health-item-sub">≈ {(total_active_minutes / 60):.1f} hours of brisk walking</div>
                    </div>
                </div>
                <div class="health-item">
                    <div class="health-item-icon">🪜</div>
                    <div class="health-item-body">
                        <div class="health-item-value">{total_floors:,.0f}</div>
                        <div class="health-item-label">Floors equivalent</div>
                        <div class="health-item-sub">≈ {eiffel_climbs:.1f}% of an Eiffel Tower climb</div>
                    </div>
                </div>
            </div>
        </div>

        <!-- Insight cards: goal-hit rate / projection / consistency -->
        <div class="insight-grid animate delay-3">
            <div class="insight-card">
                <div class="insight-icon">✅</div>
                <div class="insight-body">
                    <div class="insight-label">Goal Hit Rate</div>
                    <div class="insight-value">{goal_hit_rate:.1f}%</div>
                    <div class="insight-sub">{person_days_hit_goal:,} of {person_days_evaluated:,} person-days reached {DAILY_GOAL:,}</div>
                </div>
            </div>
            <div class="insight-card">
                <div class="insight-icon">📈</div>
                <div class="insight-body">
                    <div class="insight-label">Projected Total</div>
                    <div class="insight-value">{projected_total:,.0f}</div>
                    <div class="insight-sub">vs goal {challenge_goal_team:,} &nbsp;·&nbsp; at current pace</div>
                </div>
            </div>
            <div class="insight-card">
                <div class="insight-icon">🧘</div>
                <div class="insight-body">
                    <div class="insight-label">Most Consistent</div>
                    <div class="insight-value" style="font-size: 18px;">{most_consistent_name}</div>
                    <div class="insight-sub">⌀ {most_consistent_avg:,.0f} steps/day, lowest variation</div>
                </div>
            </div>
        </div>

        <!-- Main Charts Row -->
        <div class="charts-grid animate delay-2">
            <div class="chart-card">
                <div class="chart-title">
                    <div class="chart-title-icon" style="background: rgba(18,65,145,0.1);">📈</div>
                    Team Daily Steps
                </div>
                <div class="tab-group">
                    <button class="tab-btn active" onclick="setDailyChart('total', this)">Team Total</button>
                    <button class="tab-btn" onclick="setDailyChart('individual', this)">Individual</button>
                    <button class="tab-btn" onclick="setDailyChart('stacked', this)">Stacked</button>
                </div>
                <div class="chart-container" style="height: 340px;">
                    <canvas id="dailyChart"></canvas>
                </div>
            </div>
            <div class="chart-card">
                <div class="chart-title">
                    <div class="chart-title-icon" style="background: rgba(255,61,113,0.1);">🏅</div>
                    Step Ranking
                </div>
                <div class="leaderboard" id="leaderboard"></div>
            </div>
        </div>

        <!-- Bottom Grid -->
        <div class="bottom-grid">
            <div class="chart-card animate delay-3">
                <div class="chart-title">
                    <div class="chart-title-icon" style="background: rgba(0,157,224,0.1);">📅</div>
                    Weekly Progress
                </div>
                <div class="chart-container" style="height: 320px;">
                    <canvas id="weeklyChart"></canvas>
                </div>
            </div>
            <div class="chart-card animate delay-4">
                <div class="chart-title">
                    <div class="chart-title-icon" style="background: rgba(0,196,140,0.1);">🎯</div>
                    Daily Goal ({DAILY_GOAL:,.0f} steps)
                </div>
                <div class="chart-container" style="height: 320px;">
                    <canvas id="goalChart"></canvas>
                </div>
            </div>
        </div>

        <!-- Heatmap -->
        <div class="chart-card animate delay-5" style="margin-bottom: 28px;">
            <div class="chart-title">
                <div class="chart-title-icon" style="background: rgba(255,169,64,0.1);">🗓️</div>
                                Last 7 Days — Activity Heatmap
            </div>
            <div class="heatmap-grid" id="heatmap"></div>
        </div>

        <!-- Weekday pattern + Individual performance -->
        <div class="bottom-grid">
            <div class="chart-card animate delay-5">
                <div class="chart-title">
                    <div class="chart-title-icon" style="background: rgba(0,196,140,0.10);">📆</div>
                    Average Steps by Day of Week
                </div>
                <div class="chart-container" style="height: 320px;">
                    <canvas id="weekdayChart"></canvas>
                </div>
            </div>
            <div class="chart-card animate delay-5">
                <div class="chart-title">
                    <div class="chart-title-icon" style="background: rgba(18,65,145,0.08);">👤</div>
                    Individual Performance — Average vs Goal
                </div>
                <div class="chart-container" style="height: 320px;">
                    <canvas id="individualChart"></canvas>
                </div>
            </div>
        </div>

        <div class="footer">
            Dashboard generated on {datetime.now().strftime('%m/%d/%Y at %H:%M')} &nbsp;•&nbsp; Step Tracking Challenge 2026 — Team 8
        </div>
    </div>

    <script>
        // Data
        const memberData = {member_data_json};
        const dateLabels = {date_labels_json};
        const shortLabels = {short_labels_json};
        const weekLabels = {week_labels_json};
        const weeklyMemberData = {weekly_member_json};
        const weekdayAvg = {weekday_avg_json};
        const weekdayLabels = {weekday_labels_json};
        const DAILY_GOAL = {DAILY_GOAL};
        // Cap elapsed days to the size of the data so we never read past
        // the end of any per-member array.
        const _maxDaily = memberData.reduce((m, x) => Math.max(m, x.daily.length), 0);
        const daysElapsed = Math.min({days_elapsed}, _maxDaily);

        // Chart.js default config
        Chart.defaults.color = '#4A5568';
        Chart.defaults.borderColor = 'rgba(226,232,240,0.8)';
        Chart.defaults.font.family = "'Inter', sans-serif";

        // ===== DAILY CHART =====
        let dailyChart;
        function setDailyChart(mode, btn) {{
            document.querySelectorAll('.tab-group .tab-btn').forEach(b => b.classList.remove('active'));
            if (btn) {{
                btn.classList.add('active');
            }} else if (typeof event !== 'undefined' && event && event.target) {{
                event.target.classList.add('active');
            }}
            if (dailyChart) dailyChart.destroy();
            const ctx = document.getElementById('dailyChart').getContext('2d');
            const activeLabels = shortLabels.slice(0, daysElapsed);

            if (mode === 'total') {{
                const totals = [];
                for (let i = 0; i < daysElapsed; i++) {{
                    let sum = 0;
                    memberData.forEach(m => sum += m.daily[i]);
                    totals.push(sum);
                }}
                const gradient = ctx.createLinearGradient(0, 0, 0, 340);
                gradient.addColorStop(0, 'rgba(18,65,145,0.25)');
                gradient.addColorStop(1, 'rgba(18,65,145,0.02)');
                dailyChart = new Chart(ctx, {{
                    type: 'line',
                    data: {{
                        labels: activeLabels,
                        datasets: [{{
                            label: 'Team Total',
                            data: totals,
                            borderColor: '#124191',
                            backgroundColor: gradient,
                            fill: true,
                            tension: 0.4,
                            pointRadius: 4,
                            pointBackgroundColor: '#124191',
                            pointBorderColor: '#fff',
                            pointBorderWidth: 2,
                            borderWidth: 3,
                        }},
                        {{
                            label: 'Daily Goal',
                            data: Array(daysElapsed).fill(DAILY_GOAL * memberData.length),
                            borderColor: 'rgba(255,61,113,0.6)',
                            borderDash: [8, 4],
                            borderWidth: 2,
                            pointRadius: 0,
                            fill: false,
                        }}]
                    }},
                    options: {{
                        responsive: true,
                        maintainAspectRatio: false,
                        plugins: {{ legend: {{ display: true, position: 'top', labels: {{ usePointStyle: true, padding: 20 }} }} }},
                        scales: {{
                            y: {{ beginAtZero: true, grid: {{ color: 'rgba(226,232,240,0.6)' }},
                                ticks: {{ callback: v => (v/1000).toFixed(0) + 'k' }} }},
                            x: {{ grid: {{ display: false }} }}
                        }},
                        interaction: {{ mode: 'index', intersect: false }}
                    }}
                }});
            }} else if (mode === 'individual') {{
                const datasets = memberData.map(m => ({{
                    label: m.name.split(' ')[0],
                    data: m.daily.slice(0, daysElapsed),
                    borderColor: m.color,
                    backgroundColor: m.color + '20',
                    tension: 0.4,
                    pointRadius: 2,
                    borderWidth: 2,
                    fill: false,
                }}));
                dailyChart = new Chart(ctx, {{
                    type: 'line',
                    data: {{ labels: activeLabels, datasets }},
                    options: {{
                        responsive: true,
                        maintainAspectRatio: false,
                        plugins: {{ legend: {{ display: true, position: 'top', labels: {{ usePointStyle: true, padding: 12, font: {{ size: 11 }} }} }} }},
                        scales: {{
                            y: {{ beginAtZero: true, grid: {{ color: 'rgba(226,232,240,0.6)' }},
                                ticks: {{ callback: v => (v/1000).toFixed(0) + 'k' }} }},
                            x: {{ grid: {{ display: false }} }}
                        }},
                        interaction: {{ mode: 'index', intersect: false }}
                    }}
                }});
            }} else {{
                const datasets = memberData.map(m => ({{
                    label: m.name.split(' ')[0],
                    data: m.daily.slice(0, daysElapsed),
                    backgroundColor: m.color + 'CC',
                    borderWidth: 0,
                    borderRadius: 2,
                }}));
                dailyChart = new Chart(ctx, {{
                    type: 'bar',
                    data: {{ labels: activeLabels, datasets }},
                    options: {{
                        responsive: true,
                        maintainAspectRatio: false,
                        plugins: {{ legend: {{ display: true, position: 'top', labels: {{ usePointStyle: true, padding: 12, font: {{ size: 11 }} }} }} }},
                        scales: {{
                            y: {{ stacked: true, beginAtZero: true, grid: {{ color: 'rgba(226,232,240,0.6)' }},
                                ticks: {{ callback: v => (v/1000).toFixed(0) + 'k' }} }},
                            x: {{ stacked: true, grid: {{ display: false }} }}
                        }},
                        interaction: {{ mode: 'index', intersect: false }}
                    }}
                }});
            }}
        }}

        // Init daily chart
        (function() {{
            const ctx = document.getElementById('dailyChart').getContext('2d');
            const totals = [];
            for (let i = 0; i < daysElapsed; i++) {{
                let sum = 0;
                memberData.forEach(m => sum += m.daily[i]);
                totals.push(sum);
            }}
            const gradient = ctx.createLinearGradient(0, 0, 0, 340);
            gradient.addColorStop(0, 'rgba(18,65,145,0.25)');
            gradient.addColorStop(1, 'rgba(18,65,145,0.02)');
            dailyChart = new Chart(ctx, {{
                type: 'line',
                data: {{
                    labels: shortLabels.slice(0, daysElapsed),
                    datasets: [{{
                        label: 'Team Total',
                        data: totals,
                        borderColor: '#124191',
                        backgroundColor: gradient,
                        fill: true,
                        tension: 0.4,
                        pointRadius: 4,
                        pointBackgroundColor: '#124191',
                        pointBorderColor: '#fff',
                        pointBorderWidth: 2,
                        borderWidth: 3,
                    }},
                    {{
                        label: 'Daily Goal',
                        data: Array(daysElapsed).fill(DAILY_GOAL * memberData.length),
                        borderColor: 'rgba(255,61,113,0.6)',
                        borderDash: [8, 4],
                        borderWidth: 2,
                        pointRadius: 0,
                        fill: false,
                    }}]
                }},
                options: {{
                    responsive: true,
                    maintainAspectRatio: false,
                    plugins: {{ legend: {{ display: true, position: 'top', labels: {{ usePointStyle: true, padding: 20 }} }} }},
                    scales: {{
                        y: {{ beginAtZero: true, grid: {{ color: 'rgba(226,232,240,0.6)' }},
                            ticks: {{ callback: v => (v/1000).toFixed(0) + 'k' }} }},
                        x: {{ grid: {{ display: false }} }}
                    }},
                    interaction: {{ mode: 'index', intersect: false }}
                }}
            }});
        }})();

        // ===== LEADERBOARD =====
        (function() {{
            const sorted = [...memberData].sort((a, b) => b.total - a.total);
            const maxSteps = sorted[0]?.total || 1;
            const lb = document.getElementById('leaderboard');
            sorted.forEach((m, i) => {{
                const cls = i === 0 ? 'gold' : i === 1 ? 'silver' : i === 2 ? 'bronze' : '';
                const rankCls = i >= 3 ? 'other' : '';
                const pct = (m.total / maxSteps * 100).toFixed(0);
                const chips = [];
                if (m.goalDays > 0) {{
                    chips.push('<span class="leader-chip">✓ ' + m.goalDays + ' goal-day' + (m.goalDays === 1 ? '' : 's') + '</span>');
                }}
                if (m.streak >= 2) {{
                    chips.push('<span class="leader-chip streak">🔥 ' + m.streak + '-day streak</span>');
                }}
                const chipsHtml = chips.length ? '<div class="leader-chips">' + chips.join('') + '</div>' : '';
                lb.innerHTML += `
                    <div class="leader-row ${{cls}}">
                        <div class="leader-rank ${{rankCls}}">${{i+1}}</div>
                        <div class="leader-info">
                            <div class="leader-name">${{m.name}}</div>
                            <div class="leader-avg">⌀ ${{m.avg.toLocaleString()}} / day</div>
                            ${{chipsHtml}}
                        </div>
                        <div class="leader-bar-wrapper">
                            <div class="leader-bar" style="width: ${{pct}}%; background: ${{m.color}};"></div>
                        </div>
                        <div class="leader-steps">${{m.total.toLocaleString()}}</div>
                    </div>`;
            }});
        }})();

        // ===== WEEKLY CHART =====
        (function() {{
            const ctx = document.getElementById('weeklyChart').getContext('2d');
            const datasets = memberData.map(m => ({{
                label: m.name.split(' ')[0],
                data: weeklyMemberData[m.name] || [],
                backgroundColor: m.color + 'DD',
                borderWidth: 0,
                borderRadius: 6,
            }}));
            new Chart(ctx, {{
                type: 'bar',
                data: {{ labels: weekLabels, datasets }},
                options: {{
                    responsive: true,
                    maintainAspectRatio: false,
                    plugins: {{ legend: {{ display: true, position: 'top', labels: {{ usePointStyle: true, padding: 12, font: {{ size: 11 }} }} }} }},
                    scales: {{
                        y: {{ stacked: true, beginAtZero: true, grid: {{ color: 'rgba(226,232,240,0.6)' }},
                            ticks: {{ callback: v => (v/1000).toFixed(0) + 'k' }} }},
                        x: {{ stacked: true, grid: {{ display: false }} }}
                    }}
                }}
            }});
        }})();

        // ===== GOAL CHART (Radar) =====
        (function() {{
            const ctx = document.getElementById('goalChart').getContext('2d');
            const names = memberData.map(m => m.name.split(' ')[0]);
            const avgs = memberData.map(m => m.avg);
            const goals = memberData.map(() => DAILY_GOAL);

            new Chart(ctx, {{
                type: 'radar',
                data: {{
                    labels: names,
                    datasets: [
                        {{
                            label: 'Daily Average',
                            data: avgs,
                            borderColor: '#124191',
                            backgroundColor: 'rgba(18,65,145,0.12)',
                            borderWidth: 2,
                            pointBackgroundColor: '#124191',
                            pointRadius: 4,
                        }},
                        {{
                            label: 'Goal',
                            data: goals,
                            borderColor: 'rgba(255,61,113,0.6)',
                            backgroundColor: 'rgba(255,61,113,0.05)',
                            borderWidth: 2,
                            borderDash: [5, 3],
                            pointRadius: 0,
                        }}
                    ]
                }},
                options: {{
                    responsive: true,
                    maintainAspectRatio: false,
                    plugins: {{ legend: {{ position: 'top', labels: {{ usePointStyle: true, padding: 16 }} }} }},
                    scales: {{
                        r: {{
                            beginAtZero: true,
                            grid: {{ color: 'rgba(226,232,240,0.6)' }},
                            angleLines: {{ color: 'rgba(226,232,240,0.6)' }},
                            pointLabels: {{ font: {{ size: 11, weight: '600' }}, color: '#8892A4' }},
                            ticks: {{ display: false }}
                        }}
                    }}
                }}
            }});
        }})();

        // ===== HEATMAP =====
        (function() {{
            const grid = document.getElementById('heatmap');

            // Get last 7 days indices
            const last7Start = Math.max(0, daysElapsed - 7);
            const last7End = daysElapsed;
            const days7Labels = dateLabels.slice(last7Start, last7End);

            // Headers
            grid.innerHTML = '<div class="heatmap-header"></div>';
            days7Labels.forEach(d => {{
                grid.innerHTML += `<div class="heatmap-header">${{d}}</div>`;
            }});
            // Pad if less than 7
            for (let i = days7Labels.length; i < 7; i++) {{
                grid.innerHTML += `<div class="heatmap-header">—</div>`;
            }}

            // Member rows
            memberData.forEach(m => {{
                grid.innerHTML += `<div class="heatmap-name">${{m.name.split(' ')[0]}}</div>`;
                for (let i = last7Start; i < last7End; i++) {{
                    const val = m.daily[i] || 0;
                    const pct = Math.min(val / DAILY_GOAL, 1.5);
                    let bg, fg;
                    if (val === 0) {{
                        bg = 'rgba(226,232,240,0.4)';
                        fg = '#8892A4';
                    }} else if (pct < 0.5) {{
                        bg = 'rgba(255,61,113,0.2)';
                        fg = '#E0245E';
                    }} else if (pct < 0.8) {{
                        bg = 'rgba(255,169,64,0.2)';
                        fg = '#D48400';
                    }} else if (pct < 1.0) {{
                        bg = 'rgba(0,157,224,0.15)';
                        fg = '#009DE0';
                    }} else {{
                        bg = 'rgba(0,196,140,0.2)';
                        fg = '#009E73';
                    }}
                    const display = val > 0 ? (val/1000).toFixed(1) + 'k' : '—';
                    grid.innerHTML += `<div class="heatmap-cell" style="background:${{bg}}; color:${{fg}};">${{display}}</div>`;
                }}
                // Pad
                for (let i = days7Labels.length; i < 7; i++) {{
                    grid.innerHTML += `<div class="heatmap-cell" style="background:rgba(226,232,240,0.3); color:#8892A4;">—</div>`;
                }}
            }});
        }})();

        // ===== WEEKDAY PATTERN CHART =====
        (function() {{
            const ctx = document.getElementById('weekdayChart').getContext('2d');
            const gradient = ctx.createLinearGradient(0, 0, 0, 320);
            gradient.addColorStop(0, 'rgba(0,196,140,0.85)');
            gradient.addColorStop(1, 'rgba(0,157,224,0.55)');
            new Chart(ctx, {{
                type: 'bar',
                data: {{
                    labels: weekdayLabels,
                    datasets: [{{
                        label: 'Avg steps / member',
                        data: weekdayAvg,
                        backgroundColor: gradient,
                        borderRadius: 10,
                        borderWidth: 0,
                    }}, {{
                        label: 'Daily Goal',
                        type: 'line',
                        data: Array(7).fill(DAILY_GOAL),
                        borderColor: 'rgba(255,61,113,0.6)',
                        borderDash: [6, 4],
                        borderWidth: 2,
                        pointRadius: 0,
                        fill: false,
                    }}]
                }},
                options: {{
                    responsive: true,
                    maintainAspectRatio: false,
                    plugins: {{
                        legend: {{ display: true, position: 'top', labels: {{ usePointStyle: true, padding: 16, font: {{ size: 11 }} }} }},
                        tooltip: {{
                            callbacks: {{
                                label: function(ctx) {{
                                    return ctx.dataset.label + ': ' + Math.round(ctx.parsed.y).toLocaleString() + ' steps';
                                }}
                            }}
                        }}
                    }},
                    scales: {{
                        y: {{ beginAtZero: true, grid: {{ color: 'rgba(226,232,240,0.6)' }},
                            ticks: {{ callback: v => (v/1000).toFixed(1) + 'k' }} }},
                        x: {{ grid: {{ display: false }} }}
                    }}
                }}
            }});
        }})();

        // ===== INDIVIDUAL CHART =====
        (function() {{
            const ctx = document.getElementById('individualChart').getContext('2d');
            const sorted = [...memberData].sort((a, b) => b.avg - a.avg);
            const names = sorted.map(m => m.name);
            const avgs = sorted.map(m => m.avg);
            const bgColors = sorted.map(m => m.color + 'CC');
            const borderColors = sorted.map(m => m.color);

            new Chart(ctx, {{
                type: 'bar',
                data: {{
                    labels: names,
                    datasets: [
                        {{
                            label: 'Daily Average',
                            data: avgs,
                            backgroundColor: bgColors,
                            borderColor: borderColors,
                            borderWidth: 2,
                            borderRadius: 8,
                            barThickness: 36,
                        }}
                    ]
                }},
                options: {{
                    indexAxis: 'y',
                    responsive: true,
                    maintainAspectRatio: false,
                    plugins: {{
                        legend: {{ display: false }},
                        annotation: {{}}
                    }},
                    scales: {{
                        x: {{
                            beginAtZero: true,
                            grid: {{ color: 'rgba(226,232,240,0.6)' }},
                            ticks: {{ callback: v => (v/1000).toFixed(0) + 'k' }}
                        }},
                        y: {{
                            grid: {{ display: false }},
                            ticks: {{ font: {{ size: 12, weight: '600' }} }}
                        }}
                    }}
                }},
                plugins: [{{
                    afterDraw: function(chart) {{
                        const ctx = chart.ctx;
                        // Draw goal line
                        const xScale = chart.scales.x;
                        const goalX = xScale.getPixelForValue(DAILY_GOAL);
                        const yScale = chart.scales.y;
                        ctx.save();
                        ctx.strokeStyle = 'rgba(255,61,113,0.6)';
                        ctx.lineWidth = 2;
                        ctx.setLineDash([6, 4]);
                        ctx.beginPath();
                        ctx.moveTo(goalX, yScale.top);
                        ctx.lineTo(goalX, yScale.bottom);
                        ctx.stroke();
                        ctx.fillStyle = '#E0245E';
                        ctx.font = '11px Inter';
                        ctx.fillText('Goal', goalX + 4, yScale.top + 12);
                        ctx.restore();
                    }}
                }}]
            }});
        }})();

        // Number animation for KPIs
        function animateValue(id, start, end, duration) {{
            const el = document.getElementById(id);
            if (!el || end === 0) return;
            const range = end - start;
            const startTime = performance.now();
            function update(currentTime) {{
                const elapsed = currentTime - startTime;
                const progress = Math.min(elapsed / duration, 1);
                const eased = 1 - Math.pow(1 - progress, 3);
                const current = Math.floor(start + range * eased);
                el.textContent = current.toLocaleString();
                if (progress < 1) requestAnimationFrame(update);
            }}
            requestAnimationFrame(update);
        }}

        animateValue('kpi-total', 0, {int(total_team)}, 1500);
        animateValue('kpi-avg', 0, {int(avg_per_day_team)}, 1200);
    </script>

    <!-- Upload Token Modal -->
    <div class="upload-overlay" id="tokenOverlay">
        <div class="upload-modal">
            <h3>🔑 GitHub Token</h3>
            <p>Para enviar a planilha, insira seu Personal Access Token do GitHub (com permissão <strong>repo</strong>). Ele será salvo localmente no seu navegador.</p>
            <input type="password" id="tokenInput" placeholder="ghp_xxxxxxxxxxxx">
            <div class="upload-modal-btns">
                <button class="modal-btn-cancel" id="tokenCancel">Cancelar</button>
                <button class="modal-btn-primary" id="tokenSave">Salvar e Enviar</button>
            </div>
        </div>
    </div>

    <!-- Toast container -->
    <div id="uploadToast" class="upload-toast"></div>

    <script>
    (function() {{
        const OWNER = 'eronbfr';
        const REPO = 'Team-8-STC-Nokia';
        const FILE_PATH = 'step-tracking_Team8.xlsx';
        const API_BASE = 'https://api.github.com';
        const TOKEN_KEY = 'gh_pat_team8';

        const uploadBtn = document.getElementById('uploadBtn');
        const fileInput = document.getElementById('xlsxFileInput');
        const tokenOverlay = document.getElementById('tokenOverlay');
        const tokenInput = document.getElementById('tokenInput');
        const tokenCancel = document.getElementById('tokenCancel');
        const tokenSave = document.getElementById('tokenSave');
        const toastEl = document.getElementById('uploadToast');

        let pendingFile = null;

        function showToast(msg, type, duration) {{
            toastEl.textContent = msg;
            toastEl.className = 'upload-toast ' + type + ' show';
            if (duration) {{
                setTimeout(function() {{ toastEl.classList.remove('show'); }}, duration);
            }}
        }}

        function hideToast() {{
            toastEl.classList.remove('show');
        }}

        function setButtonLoading(loading) {{
            if (loading) {{
                uploadBtn.disabled = true;
                uploadBtn.innerHTML = '<span class="upload-spinner"></span> Enviando…';
            }} else {{
                uploadBtn.disabled = false;
                uploadBtn.innerHTML = '📤 Upload';
            }}
        }}

        // Click upload → open file picker
        uploadBtn.addEventListener('click', function() {{
            fileInput.value = '';
            fileInput.click();
        }});

        // File selected
        fileInput.addEventListener('change', function() {{
            if (!fileInput.files || !fileInput.files[0]) return;
            var file = fileInput.files[0];
            if (!file.name.endsWith('.xlsx')) {{
                showToast('❌ Selecione um arquivo .xlsx', 'error', 4000);
                return;
            }}
            pendingFile = file;
            var token = localStorage.getItem(TOKEN_KEY);
            if (token) {{
                doUpload(token);
            }} else {{
                tokenInput.value = '';
                tokenOverlay.classList.add('active');
                tokenInput.focus();
            }}
        }});

        // Token modal cancel
        tokenCancel.addEventListener('click', function() {{
            tokenOverlay.classList.remove('active');
            pendingFile = null;
        }});

        // Token modal save
        tokenSave.addEventListener('click', function() {{
            var token = tokenInput.value.trim();
            if (!token) {{
                tokenInput.style.borderColor = '#ef4444';
                return;
            }}
            localStorage.setItem(TOKEN_KEY, token);
            tokenOverlay.classList.remove('active');
            doUpload(token);
        }});

        // Close modal on overlay click
        tokenOverlay.addEventListener('click', function(e) {{
            if (e.target === tokenOverlay) {{
                tokenOverlay.classList.remove('active');
                pendingFile = null;
            }}
        }});

        function fileToBase64(file) {{
            return new Promise(function(resolve, reject) {{
                var reader = new FileReader();
                reader.onload = function() {{
                    var base64 = reader.result.split(',')[1];
                    resolve(base64);
                }};
                reader.onerror = reject;
                reader.readAsDataURL(file);
            }});
        }}

        async function doUpload(token) {{
            if (!pendingFile) return;
            setButtonLoading(true);
            showToast('📤 Enviando planilha…', 'info', 0);

            try {{
                // 1. Get current file SHA (if it exists)
                var sha = null;
                var getResp = await fetch(API_BASE + '/repos/' + OWNER + '/' + REPO + '/contents/' + FILE_PATH, {{
                    headers: {{ 'Authorization': 'Bearer ' + token }}
                }});
                if (getResp.ok) {{
                    var data = await getResp.json();
                    sha = data.sha;
                }} else if (getResp.status === 401 || getResp.status === 403) {{
                    localStorage.removeItem(TOKEN_KEY);
                    showToast('❌ Token inválido ou sem permissão. Tente novamente.', 'error', 5000);
                    setButtonLoading(false);
                    pendingFile = null;
                    return;
                }}

                // 2. Upload via Contents API
                var base64Content = await fileToBase64(pendingFile);
                var body = {{
                    message: 'Update step-tracking spreadsheet via dashboard',
                    content: base64Content,
                    branch: 'main'
                }};
                if (sha) body.sha = sha;

                var putResp = await fetch(API_BASE + '/repos/' + OWNER + '/' + REPO + '/contents/' + FILE_PATH, {{
                    method: 'PUT',
                    headers: {{
                        'Authorization': 'Bearer ' + token,
                        'Content-Type': 'application/json'
                    }},
                    body: JSON.stringify(body)
                }});

                if (putResp.ok) {{
                    showToast('✅ Planilha enviada! Atualizando dashboard…', 'success', 0);
                    setButtonLoading(false);
                    pendingFile = null;
                    // Poll for workflow completion then refresh
                    pollForRefresh(token);
                }} else if (putResp.status === 401 || putResp.status === 403) {{
                    localStorage.removeItem(TOKEN_KEY);
                    showToast('❌ Token inválido ou sem permissão. Tente novamente.', 'error', 5000);
                    setButtonLoading(false);
                    pendingFile = null;
                }} else if (putResp.status === 409) {{
                    showToast('⚠️ Conflito ao enviar. Recarregue a página e tente novamente.', 'error', 5000);
                    setButtonLoading(false);
                    pendingFile = null;
                }} else {{
                    var errText = await putResp.text();
                    showToast('❌ Erro ao enviar: ' + putResp.status + ' — ' + errText.slice(0, 100), 'error', 5000);
                    setButtonLoading(false);
                    pendingFile = null;
                }}
            }} catch (e) {{
                showToast('❌ Erro de rede: ' + e.message, 'error', 5000);
                setButtonLoading(false);
                pendingFile = null;
            }}
        }}

        // Poll GitHub Actions workflow until the dashboard is updated, then force refresh
        function pollForRefresh(token) {{
            var attempts = 0;
            var maxAttempts = 60; // poll for up to ~5 min
            var interval = 5000; // every 5 seconds

            var timer = setInterval(async function() {{
                attempts++;
                if (attempts > maxAttempts) {{
                    clearInterval(timer);
                    showToast('⏱️ Tempo esgotado. Recarregue manualmente.', 'info', 6000);
                    return;
                }}

                try {{
                    // Check latest "Update Dashboard" workflow runs (by workflow file)
                    var resp = await fetch(API_BASE + '/repos/' + OWNER + '/' + REPO + '/actions/workflows/update-dashboard.yml/runs?per_page=1&branch=main', {{
                        headers: {{ 'Authorization': 'Bearer ' + token }}
                    }});
                    if (!resp.ok) {{
                        // Fallback: try the generic runs endpoint
                        resp = await fetch(API_BASE + '/repos/' + OWNER + '/' + REPO + '/actions/runs?per_page=1&branch=main', {{
                            headers: {{ 'Authorization': 'Bearer ' + token }}
                        }});
                    }}
                    if (!resp.ok) {{
                        clearInterval(timer);
                        // Even without API access, try refreshing after a reasonable delay
                        setTimeout(function() {{ location.reload(); }}, 60000);
                        return;
                    }}
                    var data = await resp.json();
                    if (data.workflow_runs && data.workflow_runs.length > 0) {{
                        var run = data.workflow_runs[0];
                        if (run.status === 'completed') {{
                            clearInterval(timer);
                            showToast('🔄 Dashboard atualizado! Recarregando…', 'success', 2000);
                            setTimeout(function() {{ location.reload(); }}, 2000);
                        }}
                    }}
                }} catch (e) {{
                    // Silently continue polling
                }}
            }}, interval);
        }}
    }})();
    </script>
</body>
</html>"""

    return html


def main():
    print("📖 Lendo dados do Excel...")
    dates, members, all_zero = read_excel_data()

    if all_zero:
        print("⚠️  Dados vazios. Gerando com dados de demonstração...")
        members = generate_demo_data(dates, members)
        is_demo = True
    else:
        is_demo = False

    print("🎨 Gerando dashboard...")
    html = generate_html(dates, members, is_demo)

    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        f.write(html)

    print(f"✅ Dashboard salvo em: {OUTPUT_PATH}")
    if not os.environ.get('CI'):
        print("🌐 Abrindo no navegador...")
        webbrowser.open(f'file:///{OUTPUT_PATH.replace(os.sep, "/")}')


if __name__ == '__main__':
    main()
